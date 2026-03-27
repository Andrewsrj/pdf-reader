from __future__ import annotations

import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

from application.extraction_service import InvoiceBatchProcessor
from infrastructure.export.excel_exporter import ExcelExporter
from tests.integration.sample_expectations import (
    EXPECTED_SAMPLE_RESULTS,
    EXPECTED_SUMMARY_ROWS,
    EXPECTED_TOTAL_FILES,
    EXPECTED_TOTAL_ITEMS,
)


DEFAULT_SAMPLE_DIR = Path("tests/fixtures/pdfs")


def _resolve_sample_dir() -> Path | None:
    if env_path := os.getenv("PDF_READER_SAMPLE_DIR"):
        return Path(env_path)

    if DEFAULT_SAMPLE_DIR.exists():
        return DEFAULT_SAMPLE_DIR

    return None


@pytest.fixture(scope="module")
def sample_pdf_paths() -> list[Path]:
    sample_dir = _resolve_sample_dir()
    if sample_dir is None:
        pytest.skip(
            "Nenhum diretorio de amostras configurado. Defina PDF_READER_SAMPLE_DIR ou use tests/fixtures/pdfs."
        )

    missing_files = [
        filename
        for filename in EXPECTED_SAMPLE_RESULTS
        if not (sample_dir / filename).exists()
    ]
    if missing_files:
        pytest.skip(
            "Diretorio de amostras incompleto. Arquivos ausentes: " + ", ".join(sorted(missing_files))
        )

    return [sample_dir / filename for filename in sorted(EXPECTED_SAMPLE_RESULTS)]


@pytest.fixture(scope="module")
def extracted_batch(sample_pdf_paths: list[Path]):
    return InvoiceBatchProcessor().extract_batch(sample_pdf_paths)


def test_sample_pipeline_matches_expected_extraction(extracted_batch) -> None:
    assert extracted_batch.total_files == EXPECTED_TOTAL_FILES
    assert extracted_batch.total_items == EXPECTED_TOTAL_ITEMS
    assert extracted_batch.succeeded == EXPECTED_TOTAL_FILES
    assert extracted_batch.partial == 0
    assert extracted_batch.failed == 0

    results_by_name = {result.pdf_path.name: result for result in extracted_batch.results}

    for filename, expected in EXPECTED_SAMPLE_RESULTS.items():
        result = results_by_name[filename]
        assert result.city == expected["city"]
        assert result.document_number == expected["document_number"]
        assert result.issue_date == expected["issue_date"]
        assert [
            (item.item_description, item.item_quantity, item.item_total_value)
            for item in result.items
        ] == expected["items"]


def test_sample_pipeline_matches_expected_summary(extracted_batch) -> None:
    assert [
        (
            row.city,
            row.item_description,
            row.total_item_quantity,
            row.total_item_value,
        )
        for row in extracted_batch.summary_rows
    ] == EXPECTED_SUMMARY_ROWS


def test_sample_pipeline_exports_expected_workbook(extracted_batch, tmp_path: Path) -> None:
    output_path = ExcelExporter().export_batch(tmp_path / "sample_report.xlsx", extracted_batch)
    workbook = load_workbook(output_path)

    assert workbook.sheetnames == ["Base_Itens", "Resumo_Cidade_Item", "Erros"]
    assert workbook["Base_Itens"].max_row - 1 == EXPECTED_TOTAL_ITEMS
    assert workbook["Resumo_Cidade_Item"].max_row - 1 == len(EXPECTED_SUMMARY_ROWS) + 1
    assert workbook["Erros"].max_row == 1
    assert workbook["Resumo_Cidade_Item"][f"A{workbook['Resumo_Cidade_Item'].max_row}"].value == "TOTAL GERAL"
