from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from domain.models import (
    BatchExtractionResult,
    CityItemSummaryRow,
    ExtractedItem,
    ExtractionError,
    InvoiceExtractionResult,
)
from infrastructure.export.excel_exporter import ExcelExporter


def test_excel_exporter_writes_required_sheets_and_columns(tmp_path: Path) -> None:
    batch_result = BatchExtractionResult(
        results=[
            InvoiceExtractionResult(
                pdf_path=Path("nota_a.pdf"),
                city="FORTALEZA",
                document_number="000.000.001",
                issue_date=date(2026, 2, 27),
                items=[
                    ExtractedItem(
                        pdf_filename="nota_a.pdf",
                        extraction_status="sucesso",
                        city="FORTALEZA",
                        document_number="000.000.001",
                        issue_date=date(2026, 2, 27),
                        item_description="ITEM A",
                        item_quantity=Decimal("2.000"),
                        item_total_value=Decimal("100.50"),
                    )
                ],
            ),
            InvoiceExtractionResult(
                pdf_path=Path("nota_b.pdf"),
                errors=[
                    ExtractionError(
                        pdf_filename="nota_b.pdf",
                        error_type="parser",
                        message="Nenhum item encontrado.",
                        stage="parser",
                    )
                ],
            ),
        ],
        summary_rows=[
            CityItemSummaryRow(
                city="FORTALEZA",
                item_description="ITEM A",
                total_item_quantity=Decimal("2.000"),
                total_item_value=Decimal("100.50"),
            )
        ],
    )

    output_path = ExcelExporter().export_batch(tmp_path / "relatorio", batch_result)

    assert output_path.name == "relatorio.xlsx"
    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["Base_Itens", "Resumo_Cidade_Item", "Erros"]

    base_sheet = workbook["Base_Itens"]
    assert [cell.value for cell in base_sheet[1]] == [
        "arquivo_pdf",
        "status_extracao",
        "cidade",
        "numero_documento",
        "data_emissao",
        "item_descricao",
        "item_qtd",
        "item_valor_total",
        "observacoes",
    ]
    assert base_sheet["A2"].value == "nota_a.pdf"
    assert base_sheet["E2"].value == "2026-02-27"
    assert base_sheet["F2"].value == "ITEM A"
    assert base_sheet["G2"].value == 2
    assert base_sheet["H2"].value == 100.5

    summary_sheet = workbook["Resumo_Cidade_Item"]
    assert [cell.value for cell in summary_sheet[1]] == [
        "cidade",
        "item_descricao",
        "soma_item_qtd",
        "soma_item_valor_total",
    ]
    assert summary_sheet["A2"].value == "FORTALEZA"
    assert summary_sheet["B2"].value == "ITEM A"
    assert summary_sheet["C2"].value == 2
    assert summary_sheet["D2"].value == 100.5
    assert summary_sheet["A3"].value == "TOTAL GERAL"
    assert summary_sheet["C3"].value == 2
    assert summary_sheet["D3"].value == 100.5

    error_sheet = workbook["Erros"]
    assert [cell.value for cell in error_sheet[1]] == [
        "arquivo_pdf",
        "tipo_erro",
        "mensagem",
        "etapa",
    ]
    assert error_sheet["A2"].value == "nota_b.pdf"
    assert error_sheet["B2"].value == "parser"
    assert error_sheet["C2"].value == "Nenhum item encontrado."
    assert error_sheet["D2"].value == "parser"
